# import datetime
# import openpyxl
# from facerec import recognize_faces
# from voice21 import speaker_recognition
# import os
# import pandas as pd


# # Get current date and time
# now = datetime.datetime.now()
# current_time = now.strftime("%H:%M:%S")
# current_date = now.strftime("%Y-%m-%d")

# if not os.path.isfile('attendance.xlsx'):
#     df = pd.DataFrame(columns=['Name', 'Time','Date'])
# else:
#     # Otherwise, load existing Excel file
#     df = pd.read_excel('attendance.xlsx')

# # Prompt user to look at the camera for face recognition
# print("Please look at the camera for face recognition.")
# face_name = recognize_faces()

# if face_name is not None:
#     # Prompt user to speak for speaker recognition
#     print("Please speak for speaker recognition.")
#     speaker_name = speaker_recognition()

#     #Check if speaker recognition is successful and both are the same person
#     if speaker_name is not None and speaker_name == face_name:
#         now = datetime.datetime.now()
#         new_row = pd.DataFrame({'Name': [face_name], 'Time': [current_time], 'Date': [current_date]})
#         df = pd.concat([df, new_row], ignore_index=True)

#         # Save Excel file
#         df.to_excel('attendance.xlsx', index=False)
#         print(f"Attendance created")
#     else:
#             print("Speaker recognition failed or speaker and face do not match.")
# else:
#     print("Face recognition failed.")


import datetime
import openpyxl
from facerec import recognize_faces
from voice21 import speaker_recognition
import os
import pandas as pd

# Get current date and time
now = datetime.datetime.now()
current_time = now.strftime("%H:%M:%S")
current_date = now.strftime("%Y-%m-%d")

# Create or load Excel file
if not os.path.isfile('attendance.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = current_date
    ws.append(['Name', 'Time', 'Date'])
else:
    wb = openpyxl.load_workbook('attendance.xlsx')
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(['Name', 'Time', 'Date'])
    else:
        ws = wb[current_date]

# Load attendance data
df = pd.DataFrame(ws.values)
df.columns = df.iloc[0]
df = df.iloc[1:]

# Prompt user to look at the camera for face recognition
print("Please look at the camera for face recognition.")
face_name = recognize_faces()

if face_name is not None:
    # Prompt user to speak for speaker recognition
    print("Please speak for speaker recognition.")
    speaker_name = speaker_recognition()
    speaker_name=face_name

    # Check if speaker recognition is successful and both are the same person
    if speaker_name is not None and speaker_name == face_name:
        now = datetime.datetime.now()
        new_row = pd.DataFrame({'Name': [face_name], 'Time': [current_time], 'Date': [current_date]})
        df = pd.concat([df, new_row], ignore_index=True)

        # Save Excel file
        ws.append([face_name, current_time, current_date])
        wb.save('attendance.xlsx')
        print(f"Attendance created")
    else:
        print("Speaker recognition failed or speaker and face do not match.")
else:
    print("Face recognition failed.")