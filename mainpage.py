# import tkinter as tk
# from tkinter import *
# from PIL import ImageTk, Image
# from tkinter import messagebox
# import cv2
# import os
# import face_recognition
# import speech_recognition as sr

# class MyGUI:
#     def __init__(self, root):
#         self.root = root
#         self.root.geometry("1920x1080")
#         self.root.title("My App")
        
#         self.create_background()
#         self.create_buttons()
        
#     def create_background(self):
#         bg_image = Image.open("background.jpg")
#         bg_image = bg_image.resize((1920, 1080), Image.ANTIALIAS)
#         self.bg_image = ImageTk.PhotoImage(bg_image)
        
#         bg_label = tk.Label(self.root, image=self.bg_image)
#         bg_label.place(x=0, y=0, relwidth=1, relheight=1)

#         # Add a heading
#         self.heading = tk.Label(self.root, text="ATTENDANCE", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
#         self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
#         self.heading.place(x=720, y=140,anchor="center")
        
#     def create_buttons(self):
#         add_data_image = Image.open("add.jpg")
#         add_data_image = add_data_image.resize((400, 400), Image.ANTIALIAS)
#         self.add_data_image = ImageTk.PhotoImage(add_data_image)
#         self.add_data_label = tk.Label(self.root, text="Add Data", font=("Arial", 20,"bold"),bg="gray94",fg="black")
        
#         mark_attendance_image = Image.open("mark.jpg")
#         mark_attendance_image = mark_attendance_image.resize((400, 400), Image.ANTIALIAS)
#         self.mark_attendance_image = ImageTk.PhotoImage(mark_attendance_image)
#         self.mark_attendance_label = tk.Label(self.root, text="Mark Attendance", font=("Arial", 20,"bold"),bg="gray94",fg="black")
        
#         check_attendance_image = Image.open("check.jpg")
#         check_attendance_image = check_attendance_image.resize((400, 400), Image.ANTIALIAS)
#         self.check_attendance_image = ImageTk.PhotoImage(check_attendance_image)
#         self.check_attendance_label = tk.Label(self.root, text="Check Attendance", font=("Arial", 20,"bold"),bg="gray94",fg="black")
        
#         self.add_data_button = tk.Button(self.root, image=self.add_data_image, bd=0,command=self.add_data)
#         self.mark_attendance_button = tk.Button(self.root, image=self.mark_attendance_image, bd=0,command=self.markatt)
#         self.check_attendance_button = tk.Button(self.root, image=self.check_attendance_image, bd=0,command=self.checkatt)
        
#         self.add_data_button.place(relx=0.05, rely=0.35)
#         self.mark_attendance_button.place(relx=0.35, rely=0.35)
#         self.check_attendance_button.place(relx=0.65, rely=0.35)
#         self.add_data_label.place(relx=0.19, rely=0.83, anchor="center")
#         self.mark_attendance_label.place(relx=0.496, rely=0.83, anchor="center")
#         self.check_attendance_label.place(relx=0.795, rely=0.83, anchor="center")

#     def add_data(self):
#         # tk.messagebox.showinfo("Add Data","Feature Coming Soon")
#         # self.root.destroy()

#         new_frame = tk.Frame(self.root, width=1920, bg='gray94', height=1080)
#         new_frame.place(x=0, y=0)
#         self.heading = tk.Label(new_frame, text="ADD/UPDATE DATA", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
#         self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
#         self.heading.place(x=720, y=140,anchor="center")

#         def capture_image():
#             # initialize the camera
#             cap = cv2.VideoCapture(0)
            
#             # capture the image
#             ret, frame = cap.read()
            
#             # save the image
#             image_name = "captured_image.jpg"
#             cv2.imwrite(image_name, frame)
            
#             # release the camera
#             cap.release()
            
#             # display the image in the UI
#             captured_image = Image.open(image_name)
#             captured_image = captured_image.resize((400, 400), Image.ANTIALIAS)
#             captured_image_tk = ImageTk.PhotoImage(captured_image)
#             captured_image_label = tk.Label(new_frame, image=captured_image_tk)
#             captured_image_label.image = captured_image_tk
#             captured_image_label.place(relx=0.13, rely=0.48, anchor="center")
#             tk.messagebox.showinfo("Image Captured", "Image captured and saved")
        
#         def record_voice():
#             tk.messagebox.showinfo("Voice Recorded","Feature Coming Soon")

#         # create and place the Capture Image button
#         capture_image_button = tk.Button(new_frame, text="Capture Image",command=capture_image(),bg="black",fg="Red")
#         capture_image_button.place(relx=0.086, rely=0.23)

#         # create and place the Retake Image button
#         retake_image_button = tk.Button(new_frame, text="Retake Image",command=capture_image(),bg="black",fg="Red")
#         retake_image_button.place(relx=0.085, rely=0.7)

#         # create and place the Record Voice button
#         record_voice_button = tk.Button(new_frame, text="Record Voice",command=record_voice(),bg="black",fg="red")
#         record_voice_button.place(relx=0.35, rely=0.23)

#         # create and place the Return button
#         return_button = tk.Button(new_frame, text="Return", command=new_frame.destroy)
#         return_button.place(relx=0.63, rely=0.23)

#     def markatt(self):
#         # tk.messagebox.showinfo("Mark Attendance","Feature Coming Soon")
#         # self.root.destroy()
#         mark_frame = tk.Frame(self.root, width=1920, bg='gray94', height=1080)
#         mark_frame.place(x=0, y=0)
#         self.heading = tk.Label(mark_frame, text="MARK ATTENDANCE", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
#         self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
#         self.heading.place(x=720, y=140,anchor="center")

#         start_button = tk.Button(mark_frame, text="START", font=("Arial", 30), bg="black", fg="red")
#         start_button.grid(row=2, column=1, padx=10, pady=10)
#         start_button.place(relx=0.35, rely=0.45, anchor="center")

#         def check():
#             image_dir = 'images'
#             known_faces = []
#             known_names = []
#             for filename in os.listdir(image_dir):
#                 image_path = os.path.join(image_dir, filename)
#                 image = face_recognition.load_image_file(image_path)
#                 encoding = face_recognition.face_encodings(image)[0]
#                 name = os.path.splitext(filename)[0]
#                 known_faces.append(encoding)
#                 known_names.append(name)

#             # Capture image and detect face
#             cap = cv2.VideoCapture(0)
#             ret, frame = cap.read()
#             face_locations = face_recognition.face_locations(frame)
#             if not face_locations:
#                 messagebox.showerror("Error", "No face detected")
#                 return
#             face_encodings = face_recognition.face_encodings(frame, face_locations)
#             face_encoding = face_encodings[0]

#             # Recognize face
#             matches = face_recognition.compare_faces(known_faces, face_encoding)
#             if True not in matches:
#                 messagebox.showerror("Error", "IMPOSTER")
#                 return
#             name = known_names[matches.index(True)]

#             # Prompt user to speak and recognize voice
#             r = sr.Recognizer()
#             with sr.Microphone() as source:
#                 messagebox.showinfo("Speak Now", "Speak now for 3 seconds")
#                 audio = r.record(source, duration=3)
#                 try:
#                     text = r.recognize_google(audio)
#                 except sr.UnknownValueError:
#                     messagebox.showerror("Error", "Voice not recognized")
#                     return

#             # Check if voice is present in directory
#             voice_dir = 'voices'
#             voice_files = os.listdir(voice_dir)
#             if name+'.wav' not in voice_files:
#                 messagebox.showerror("Error", "Voice not in directory")
#                 return

#             # Recognize voice
#             voice_path = os.path.join(voice_dir, name+'.wav')
#             with sr.AudioFile(voice_path) as source:
#                 audio = r.record(source)
#                 try:
#                     text = r.recognize_google(audio)
#                 except sr.UnknownValueError:
#                     messagebox.showerror("Error", "Voice not recognized")
#                     return

#             messagebox.showinfo("Attendance Marked", "Attendance marked for "+name)

#             cap.release()
        
#     def checkatt(self):
#         tk.messagebox.showinfo("Check Attendance","Feature Coming Soon")
#         # self.root.destroy()

# def create_gui():
#     root = tk.Tk()
#     my_gui = MyGUI(root)
#     root.mainloop()

# create_gui()

import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from tkinter import messagebox
import cv2
import pyaudio
import wave
import datetime
import os
import pandas as pd
from facerec import recognize_faces
from voice21 import speaker_recognition
import openpyxl

class MyGUI:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1920x1080")
        self.root.title("My App")
        
        self.create_background()
        self.create_buttons()
        
    def create_background(self):
        bg_image = Image.open("background.jpg")
        bg_image = bg_image.resize((1920, 1080), Image.LANCZOS)
        self.bg_image = ImageTk.PhotoImage(bg_image)
        
        bg_label = tk.Label(self.root, image=self.bg_image)
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)

        # Add a heading
        self.heading = tk.Label(self.root, text="ATTENDANCE", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
        self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
        self.heading.place(x=720, y=140,anchor="center")
        
    def create_buttons(self):
        add_data_image = Image.open("add.jpg")
        add_data_image = add_data_image.resize((400, 400), Image.LANCZOS)
        self.add_data_image = ImageTk.PhotoImage(add_data_image)
        self.add_data_label = tk.Label(self.root, text="Add Data", font=("Arial", 20,"bold"),bg="gray94",fg="black")
        
        mark_attendance_image = Image.open("mark.jpg")
        mark_attendance_image = mark_attendance_image.resize((400, 400), Image.LANCZOS)
        self.mark_attendance_image = ImageTk.PhotoImage(mark_attendance_image)
        self.mark_attendance_label = tk.Label(self.root, text="Mark Attendance", font=("Arial", 20,"bold"),bg="gray94",fg="black")
        
        check_attendance_image = Image.open("check.jpg")
        check_attendance_image = check_attendance_image.resize((400, 400), Image.LANCZOS)
        self.check_attendance_image = ImageTk.PhotoImage(check_attendance_image)
        self.check_attendance_label = tk.Label(self.root, text="Check Attendance", font=("Arial", 20,"bold"),bg="gray94",fg="black")
        
        self.add_data_button = tk.Button(self.root, image=self.add_data_image, bd=0,command=self.add_data)
        self.mark_attendance_button = tk.Button(self.root, image=self.mark_attendance_image, bd=0,command=self.markatt)
        self.check_attendance_button = tk.Button(self.root, image=self.check_attendance_image, bd=0,command=self.checkatt)
        
        self.add_data_button.place(relx=0.05, rely=0.35)
        self.mark_attendance_button.place(relx=0.35, rely=0.35)
        self.check_attendance_button.place(relx=0.65, rely=0.35)
        self.add_data_label.place(relx=0.19, rely=0.83, anchor="center")
        self.mark_attendance_label.place(relx=0.496, rely=0.83, anchor="center")
        self.check_attendance_label.place(relx=0.795, rely=0.83, anchor="center")

    def add_data(self):
        # tk.messagebox.showinfo("Add Data","Feature Coming Soon")
        # self.root.destroy()

        new_frame = tk.Frame(self.root, width=1920, bg='gray94', height=1080)
        new_frame.place(x=0, y=0)
        self.heading = tk.Label(new_frame, text="ADD/UPDATE DATA", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
        self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
        self.heading.place(x=720, y=140,anchor="center")

        # name_label = tk.Label(new_frame, text="Username:", font=("Arial", 20), fg="Black", bg="gray94")
        # name_entry = tk.Entry(new_frame, font=("Arial", 20), bg="gainsboro", fg="Black")
        # usn_label = tk.Label(new_frame, text="Password:", font=("Arial", 20), fg="Black", bg="gray94")
        # usn_entry = tk.Entry(new_frame, font=("Arial", 20), bg="gainsboro", fg="Black")
        # submit_button = tk.Button(new_frame, text="Login", font=("Arial", 30), bg="Orange", fg="black")

        # name_label.grid(row=0, column=0, padx=10, pady=10)
        # name_entry.grid(row=0, column=1, padx=10, pady=10)
        # usn_label.grid(row=1, column=0, padx=10, pady=10)
        # usn_entry.grid(row=1, column=1, padx=10, pady=10)
        # submit_button.grid(row=2, column=1, padx=10, pady=10)

        # name_label.place(relx=0.65, rely=0.45, anchor="center")
        # name_entry.place(relx=0.65, rely=0.49, anchor="center")
        # usn_label.place(relx=0.65, rely=0.53, anchor="center")
        # usn_entry.place(relx=0.65, rely=0.57, anchor="center")
        # submit_button.place(relx=0.65, rely=0.62, anchor="center")

        def capture_image():
            # initialize the camera
            cap = cv2.VideoCapture(0)
            
            # capture the image
            ret, frame = cap.read()
            
            # save the image
            image_name = name_entry.get()+".jpg"
            cv2.imwrite(image_name, frame)
            
            # release the camera
            cap.release()
            
            # display the image in the UI
            captured_image = Image.open(image_name)
            captured_image = captured_image.resize((400, 400), Image.LANCZOS)
            captured_image_tk = ImageTk.PhotoImage(captured_image)
            captured_image_label = tk.Label(new_frame, image=captured_image_tk)
            captured_image_label.image = captured_image_tk
            captured_image_label.place(relx=0.13, rely=0.48, anchor="center")
            tk.messagebox.showinfo("Image Captured", "Image captured and saved")
        
        def record_voice():
            FORMAT = pyaudio.paInt16
            CHANNELS = 1
            RATE = 44100
            CHUNK = 1024
            p = pyaudio.PyAudio()
            stream = p.open(format=FORMAT, channels=CHANNELS, rate=RATE, input=True, frames_per_buffer=CHUNK)
            frames = []
            for i in range(int(RATE / CHUNK * 5)):
                data = stream.read(CHUNK)
                frames.append(data)
            stream.stop_stream()
            stream.close()
            p.terminate()

            # Save audio to file
            filename = name_entry.get()+".wav"
            with wave.open(filename, 'wb') as wf:
                wf.setnchannels(CHANNELS)
                wf.setsampwidth(p.get_sample_size(FORMAT))
                wf.setframerate(RATE)
                wf.writeframes(b''.join(frames))
            
            tk.messagebox.showinfo("Voice Recorded", f"Audio recorded and saved to {filename}")

        # create and place the Capture Image button
        capture_image_button = tk.Button(new_frame, text="Capture Image",command=capture_image,bg="black",fg="Red")
        capture_image_button.place(relx=0.086, rely=0.23)

        # create and place the Retake Image button
        retake_image_button = tk.Button(new_frame, text="Retake Image",command=capture_image,bg="black",fg="Red")
        retake_image_button.place(relx=0.085, rely=0.7)

        # create and place the Record Voice button
        record_voice_button = tk.Button(new_frame, text="Record Voice",command=record_voice,bg="black",fg="red")
        record_voice_button.place(relx=0.35, rely=0.23)

        # create and place the Return button
        return_button = tk.Button(new_frame, text="Return", command=new_frame.destroy)
        return_button.place(relx=0.63, rely=0.23)

        name_label = tk.Label(new_frame, text="Username:", font=("Arial", 20), fg="Black", bg="White")
        name_entry = tk.Entry(new_frame, font=("Arial", 20), bg="gainsboro", fg="Black")

        name_label.place(relx=0.55, rely=0.6, anchor="center")
        name_entry.place(relx=0.55, rely=0.65, anchor="center")

    def markatt(self):
        new_frame1 = tk.Frame(self.root, width=1920, bg='gray94', height=1080)
        new_frame1.place(x=0, y=0)
        self.heading = tk.Label(new_frame1, text="MARK ATTENDANCE", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
        self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
        self.heading.place(x=720, y=140,anchor="center")

        def mark_attend():
            # Get current date and time
            now = datetime.datetime.now()
            current_time = now.strftime("%H:%M:%S")
            current_date = now.strftime("%Y-%m-%d")

            # Create or load Excel file
            if not os.path.isfile(subject_entry.get()+".xlsx"):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_date
                ws.append(['Name', 'Time', 'Date'])
            else:
                wb = openpyxl.load_workbook(subject_entry.get()+".xlsx")
                if current_date not in wb.sheetnames:
                    ws = wb.create_sheet(current_date)
                    ws.append(['Name', 'Time', 'Date'])
                else:
                    ws = wb[current_date]

            # Load attendance data
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]
            df = df.iloc[1:]

            # Prompt user to look at the camera for face recognition
            tk.messagebox.showinfo("Please look at the camera for face recognition.","Please look at the camera for face recognition.")
            face_name = recognize_faces()

            if face_name is not None:
                # Prompt user to speak for speaker recognition
                tk.messagebox.showinfo("Please speak for speaker recognition.","Please speak for speaker recognition.")
                speaker_name = speaker_recognition()

                # Check if speaker recognition is successful and both are the same person
                if speaker_name is not None and speaker_name == face_name:
                    now = datetime.datetime.now()
                    new_row = pd.DataFrame({'Name': [face_name], 'Time': [current_time], 'Date': [current_date]})
                    df = pd.concat([df, new_row], ignore_index=True)

                    # Save Excel file
                    ws.append([face_name, current_time, current_date])
                    wb.save(subject_entry.get()+".xlsx")
                    # tk.messagebox.showinfo("Taking speech","Taking speech")
                    tk.messagebox.showinfo("Attendance created","Attendance created")
                else:
                    tk.messagebox.showinfo("Speaker recognition failed or speaker and face do not match.")
            else:
                tk.messagebox.showinfo("Face recognition failed.")

        START_button = tk.Button(new_frame1, text="START", font=("Arial", 30), fg="red", bg="black", command=mark_attend)
        START_button.place(relx=0.35, rely=0.4, anchor="center")
        
        subject_label = tk.Label(new_frame1, text="Subject Name:", font=("Arial", 20), fg="Black", bg="White")
        subject_entry = tk.Entry(new_frame1, font=("Arial", 20), bg="gainsboro", fg="Black")

        subject_label.place(relx=0.35, rely=0.6, anchor="center")
        subject_entry.place(relx=0.35, rely=0.65, anchor="center")

        return_button = tk.Button(new_frame1, text="Return", command=new_frame1.destroy)
        return_button.place(relx=0.35, rely=0.5, anchor="center")

    def checkatt(self):
        new_frame2 = tk.Frame(self.root, width=1920, bg='gray94', height=1080)
        new_frame2.place(x=0, y=0)
        self.heading = tk.Label(new_frame2, text="CHECK ATTENDANCE", font=("Josefin Sans",70,"bold"), bg="gray94", fg="black")
        self.heading.grid(row=0, column=0, columnspan=3, padx=0, pady=0)
        self.heading.place(x=720, y=140,anchor="center")

        def check_attend():
            # Read the attendance Excel sheet into a pandas dataframe
            df = pd.read_excel(subject_entry.get()+".xlsx", sheet_name=None)

            # Prompt the user to enter the student name
            student = student_entry.get()

            # Filter the dataframe to include only rows where the student name matches
            attendance_count = 0
            for sheet_name in df.keys():
                sheet_attendance = df[sheet_name][df[sheet_name]["Name"] == student]
                attendance_count += sheet_attendance.shape[0]

            # Calculate the total number of classes held (assuming one sheet per class)
            classes_held = len(df.keys())

            # Calculate the attendance percentage
            attendance_percentage = (attendance_count / classes_held) * 100

            # Print the attendance percentage
            tk.messagebox.showinfo(f"{student} attended {attendance_count} out of {classes_held} classes ({attendance_percentage:.2f}% attendance)",f"{student} attended {attendance_count} out of {classes_held} classes ({attendance_percentage:.2f}% attendance)")

        START_button = tk.Button(new_frame2, text="CHECK", font=("Arial", 30), fg="red", bg="black", command=check_attend)
        START_button.place(relx=0.35, rely=0.4, anchor="center")
        subject_label = tk.Label(new_frame2, text="Subject Name:", font=("Arial", 20), fg="Black", bg="White")
        subject_entry = tk.Entry(new_frame2, font=("Arial", 20), bg="gainsboro", fg="Black")
        student_label = tk.Label(new_frame2, text="Student Name:", font=("Arial", 20), fg="Black", bg="White")
        student_entry = tk.Entry(new_frame2, font=("Arial", 20), bg="gainsboro", fg="Black")

        subject_label.place(relx=0.35, rely=0.6, anchor="center")
        subject_entry.place(relx=0.35, rely=0.65, anchor="center")
        student_label.place(relx=0.35, rely=0.70, anchor="center")
        student_entry.place(relx=0.35, rely=0.75, anchor="center")

        return_button = tk.Button(new_frame2, text="Return", command=new_frame2.destroy)
        return_button.place(relx=0.35, rely=0.5, anchor="center")

def create_gui():
    root = tk.Tk()
    my_gui = MyGUI(root)
    root.mainloop()

create_gui()